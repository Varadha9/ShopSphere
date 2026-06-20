from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

def header_fill(hex_color): return PatternFill("solid", fgColor=hex_color)
def thin_border():
    s = Side(style="thin", color="BBBBBB")
    return Border(left=s, right=s, top=s, bottom=s)

PASS_FILL   = PatternFill("solid", fgColor="D9F2E6")
FAIL_FILL   = PatternFill("solid", fgColor="FCE4D6")
SKIP_FILL   = PatternFill("solid", fgColor="F2F2F2")
PEND_FILL   = PatternFill("solid", fgColor="FFF2CC")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT  = Font(bold=True, size=13, color="1F3864")
CELL_FONT   = Font(size=10)

HEADERS = ["TC ID", "Module", "Test Case Title", "Preconditions",
           "Test Steps", "Test Data", "Expected Result", "Actual Result",
           "Status", "Priority", "Type"]

COL_WIDTHS = [9, 14, 38, 30, 55, 36, 38, 22, 10, 10, 14]

STATUS_FILL = {"PASS": PASS_FILL, "FAIL": FAIL_FILL,
               "SKIP": SKIP_FILL,  "": PEND_FILL}

def make_sheet(wb, title, color, test_cases, first=False):
    ws = wb.active if first else wb.create_sheet(title)
    ws.title = title
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    ws.merge_cells("A1:K1")
    ws["A1"] = f"BookSphere — {title}  ({len(test_cases)} Test Cases)"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].fill = PatternFill("solid", fgColor="EBF3FB")
    ws.row_dimensions[1].height = 28

    for col, (h, w) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = header_fill(color)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border()
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[2].height = 22

    for r, tc in enumerate(test_cases, 3):
        ws.row_dimensions[r].height = 58
        for c, val in enumerate(tc, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = CELL_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = thin_border()
            if c == 9:
                cell.fill = STATUS_FILL.get(val, PEND_FILL)
    return ws

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — Login Tests  (14 test cases)
# ═══════════════════════════════════════════════════════════════════════════════
login_tcs = [
    ["TC_L_01","Login","Valid email + password login",
     "User on login page\nInternet connected",
     "1. Open https://booksphere-dun.vercel.app\n2. Enter email: varadmandhare924@gmail.com\n3. Enter password: Varad@999\n4. Click 'Sign In →'",
     "Email: varadmandhare924@gmail.com\nPassword: Varad@999",
     "User redirected to catalog/home page",
     "Catalog page loaded — .product-card visible after login","PASS","P1","Smoke"],

    ["TC_L_02","Login","Wrong password shows lf-error-msg",
     "User on login page",
     "1. Enter email: varadmandhare924@gmail.com\n2. Enter password: wrongpass\n3. Click 'Sign In →'",
     "Email: varadmandhare924@gmail.com\nPassword: wrongpass",
     "Error message (class: lf-error-msg) appears below form",
     "lf-error-msg visible for wrong password","PASS","P1","Regression"],

    ["TC_L_03","Login","Empty email — inline field validation",
     "User on login page",
     "1. Leave email blank\n2. Enter password: Varad@999\n3. Click 'Sign In →'",
     "Email: (empty)\nPassword: Varad@999",
     "'Enter a valid email' error under email field (lf-err-msg)",
     "lf-err-msg inline validation shown for empty email","PASS","P1","Regression"],

    ["TC_L_04","Login","Empty password — inline field validation",
     "User on login page",
     "1. Enter email: varadmandhare924@gmail.com\n2. Leave password blank\n3. Click 'Sign In →'",
     "Email: varadmandhare924@gmail.com\nPassword: (empty)",
     "'Password is required' shown under password field",
     "Stays on login page for empty password","PASS","P1","Regression"],

    ["TC_L_05","Login","Invalid email format rejected",
     "User on login page",
     "1. Enter email: notanemail\n2. Enter password: Varad@999\n3. Click 'Sign In →'",
     "Email: notanemail\nPassword: Varad@999",
     "'Enter a valid email' shown — form does not submit",
     "Stays on login page for invalid email format","PASS","P2","Regression"],

    ["TC_L_06","Login","Password show/hide toggle",
     "User on login page",
     "1. Enter password: Varad@999\n2. Click 👁️ button\n3. Verify type=text\n4. Click 🙈 button\n5. Verify type=password",
     "Password: Varad@999",
     "Input type toggles between text and password",
     "Input type toggled password ↔ text correctly","PASS","P3","Functional"],

    ["TC_L_07","Login","Switch to Sign Up mode shows Name field",
     "User on login page in Sign In mode",
     "1. Click 'Sign Up' toggle button\n2. Verify Full Name field appears",
     "N/A",
     "Form switches to register mode, Full Name input visible",
     "Full Name field appeared after Sign Up toggle","PASS","P2","Functional"],

    ["TC_L_08","Login","Forgot password sends reset email",
     "User on login page",
     "1. Enter email: varadmandhare924@gmail.com\n2. Click 'Forgot password?'\n3. Check success message",
     "Email: varadmandhare924@gmail.com",
     "'✓ Reset link sent to varadmandhare924@gmail.com' visible",
     "lf-status lf-success visible after forgot password","PASS","P2","Functional"],

    ["TC_L_09","Login","Session persists after page refresh",
     "User already logged in",
     "1. Login with valid credentials\n2. Refresh browser (F5)\n3. Verify still on catalog",
     "Email: varadmandhare924@gmail.com\nPassword: Varad@999",
     "User stays on catalog — not redirected to login","","","P1","Regression"],

    ["TC_L_10","Login","Both fields empty — both errors shown",
     "User on login page",
     "1. Leave email and password both blank\n2. Click 'Sign In →'",
     "Email: (empty)\nPassword: (empty)",
     "Validation errors shown under both email and password fields","","","P1","Regression"],

    ["TC_L_11","Login","SQL injection in email field is rejected",
     "User on login page",
     "1. Enter email: ' OR '1'='1\n2. Enter password: Varad@999\n3. Click 'Sign In →'",
     "Email: ' OR '1'='1\nPassword: Varad@999",
     "Validation error shown — login does not succeed, no crash","","","P1","Security"],

    ["TC_L_12","Login","Very long email string handled gracefully",
     "User on login page",
     "1. Enter 300-character string as email\n2. Click 'Sign In →'",
     "Email: aaa...@bbb.com (300 chars)\nPassword: Varad@999",
     "Form validates / shows error — no crash or unhandled exception","","","P2","Boundary"],

    ["TC_L_13","Login","Register new account with valid data",
     "User on login page in Sign Up mode",
     "1. Click 'Sign Up' toggle\n2. Enter name: Test User\n3. Enter email: newtestuser99@gmail.com\n4. Enter password: Test@1234\n5. Click 'Create Account →'",
     "Name: Test User\nEmail: newtestuser99@gmail.com\nPassword: Test@1234",
     "Success message shown: 'Account created! Signing you in…'","","","P2","Functional"],

    ["TC_L_14","Login","Register with weak password < 6 chars",
     "User in Sign Up mode",
     "1. Click 'Sign Up' toggle\n2. Enter name, email\n3. Enter password: 123\n4. Click 'Create Account →'",
     "Password: 123 (3 chars)",
     "'Minimum 6 characters' error shown under password field","","","P2","Regression"],
]

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — Catalog Tests  (12 test cases)
# ═══════════════════════════════════════════════════════════════════════════════
catalog_tcs = [
    ["TC_C_01","Catalog","All 12 books load on catalog page",
     "User logged in",
     "1. Login successfully\n2. Navigate to catalog page\n3. Count product-card elements",
     "Valid login credentials",
     "12 book cards rendered on page",
     "12 .product-card elements rendered on catalog","PASS","P1","Smoke"],

    ["TC_C_02","Catalog","Search by book title — Atomic Habits",
     "User on catalog page",
     "1. Type 'Atomic' in search input\n2. Observe filtered results",
     "Search query: Atomic",
     "≥1 result shown, fewer cards than total 12",
     "Search 'Atomic' returned ≥1 result ≤12 total","PASS","P1","Regression"],

    ["TC_C_03","Catalog","Search by author name — James Clear",
     "User on catalog page",
     "1. Type 'James Clear' in search input",
     "Search query: James Clear",
     "Books by James Clear shown in results","","","P2","Regression"],

    ["TC_C_04","Catalog","Search by tag — 'classic'",
     "User on catalog page",
     "1. Type 'classic' in search input\n2. Check results",
     "Search query: classic",
     "Books tagged 'classic' are shown","","","P2","Regression"],

    ["TC_C_05","Catalog","Search returns empty state for garbage input",
     "User on catalog page",
     "1. Type 'xyzxyzxyz' in search input",
     "Search query: xyzxyzxyz",
     "No books shown, empty state displayed — no crash","","","P2","Regression"],

    ["TC_C_06","Catalog","Clear search restores full catalog",
     "User on catalog page, search active",
     "1. Type 'Atomic' in search\n2. Clear the search input\n3. Count books again",
     "Search query cleared",
     "All 12 books shown again after clearing search","","","P2","Regression"],

    ["TC_C_07","Catalog","Genre tree is visible and rendered",
     "User on catalog page",
     "1. Check for genre tree nodes (.tree-node)",
     "N/A",
     "Genre tree nodes visible on the page",
     ".cat-tree element visible on catalog page","PASS","P1","Smoke"],

    ["TC_C_08","Catalog","Click Fiction genre filters books",
     "Genre tree visible",
     "1. Click 'Fiction' genre node\n2. Observe book cards",
     "Genre: Fiction",
     "Only Fiction books displayed","","","P2","Regression"],

    ["TC_C_09","Catalog","Price sort — Low to High",
     "User on catalog page",
     "1. Select 'Price: Low to High' sort\n2. Read first and last book price",
     "Sort: price ascending",
     "First book has lowest price, last has highest","","","P2","Regression"],

    ["TC_C_10","Catalog","Price sort — High to Low",
     "User on catalog page",
     "1. Select 'Price: High to Low' sort\n2. Read first and last book price",
     "Sort: price descending",
     "First book has highest price, last has lowest","","","P2","Regression"],

    ["TC_C_11","Catalog","Book card shows title, author, price",
     "User on catalog page",
     "1. Look at any book card\n2. Verify title, author and price are visible",
     "N/A",
     "Each card displays title, author name and price (₹)","","","P2","Functional"],

    ["TC_C_12","Catalog","Add to Wishlist from catalog",
     "User logged in, on catalog page",
     "1. Click wishlist/heart icon on any book\n2. Navigate to Wishlist page",
     "N/A",
     "Book appears in Wishlist page (HashSet — no duplicates)","","","P2","Regression"],
]

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 3 — Cart Tests  (12 test cases)
# ═══════════════════════════════════════════════════════════════════════════════
cart_tcs = [
    ["TC_CA_01","Cart","Add book to cart from catalog",
     "User logged in, on catalog page",
     "1. Click 'Add to Cart' on first book\n2. Navigate to /cart",
     "Any book card",
     "Cart page shows ≥1 item",
     "Cart page shows ≥1 .cart-item-row after add","PASS","P1","Smoke"],

    ["TC_CA_02","Cart","Cart item count updates in navbar",
     "User on catalog page",
     "1. Note cart count in navbar\n2. Add a book to cart\n3. Check navbar count again",
     "N/A",
     "Navbar cart counter increments by 1","","","P1","Regression"],

    ["TC_CA_03","Cart","Add same book twice — quantity increases",
     "User on catalog page",
     "1. Add the same book to cart twice\n2. Go to cart\n3. Check qty for that book",
     "Same book added twice",
     "Quantity shows 2 for that book (not two separate rows)","","","P2","Regression"],

    ["TC_CA_04","Cart","Remove item from cart",
     "Cart has ≥1 item",
     "1. Go to /cart\n2. Click remove (✕) on first item\n3. Verify item gone",
     "N/A",
     "Item removed, cart count decreases by 1","","","P1","Regression"],

    ["TC_CA_05","Cart","Undo remove restores item — Stack DSA",
     "Cart has ≥1 item",
     "1. Remove an item\n2. Click 'Undo' button\n3. Check item count",
     "N/A",
     "Removed item restored — count back to original (Stack LIFO)",
     "Remove + Undo — cart count restored to original","PASS","P1","Regression"],

    ["TC_CA_06","Cart","Undo button disappears after cart empty",
     "Cart has 1 item",
     "1. Remove the only item\n2. Undo it\n3. Remove again — verify no undo visible after fresh load",
     "N/A",
     "Undo only available when stack has items","","","P3","Functional"],

    ["TC_CA_07","Cart","Apply coupon BOOK30 — min spend ₹200",
     "Cart total ≥ ₹200",
     "1. Enter coupon: BOOK30\n2. Click Apply\n3. Check discount row",
     "Coupon: BOOK30\nMin spend: ₹200\nDiscount: ₹30",
     "₹30 discount shown, total reduced by ₹30",
     ".discount-result visible after Apply Best Discount clicked","PASS","P1","Regression"],

    ["TC_CA_08","Cart","Apply coupon BOOK80 — min spend ₹500",
     "Cart total ≥ ₹500",
     "1. Enter coupon: BOOK80\n2. Click Apply",
     "Coupon: BOOK80\nMin spend: ₹500\nDiscount: ₹80",
     "₹80 discount applied","","","P1","Regression"],

    ["TC_CA_09","Cart","Apply coupon BOOK150 — min spend ₹900",
     "Cart total ≥ ₹900",
     "1. Enter coupon: BOOK150\n2. Click Apply",
     "Coupon: BOOK150\nMin spend: ₹900\nDiscount: ₹150",
     "₹150 discount applied","","","P2","Regression"],

    ["TC_CA_10","Cart","Apply invalid coupon shows no discount",
     "Cart has items",
     "1. Enter coupon: FAKECODE\n2. Click Apply",
     "Coupon: FAKECODE",
     "No discount applied, error or empty state shown — no crash","","","P2","Regression"],

    ["TC_CA_11","Cart","Discount engine — DP picks best combination",
     "Cart total ≥ ₹900, all coupons available",
     "1. Add books worth ≥ ₹900\n2. Let DP engine auto-select best coupons\n3. Verify discount is optimal",
     "Total ≥ ₹900\nBest combo: BOOK30 + BOOK80 + BOOK150 = ₹260",
     "DP selects maximum discount without exceeding budget (0/1 Knapsack)","","","P1","Regression"],

    ["TC_CA_12","Cart","Place standard order clears cart",
     "Cart has items, user logged in",
     "1. Click 'Place Order'\n2. Go back to cart",
     "N/A",
     "Cart is empty after order placed, order appears in Orders page","","","P1","Regression"],
]

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 4 — Orders Tests  (7 test cases)
# ═══════════════════════════════════════════════════════════════════════════════
orders_tcs = [
    ["TC_O_01","Orders","Orders page loads with placed orders",
     "User has placed ≥1 order",
     "1. Navigate to /orders\n2. Check order cards visible",
     "N/A",
     "Order cards rendered on the page","","","P1","Smoke"],

    ["TC_O_02","Orders","Premium order appears first — Priority Queue DSA",
     "User is premium, has both premium and standard orders",
     "1. Place premium order (priority=1)\n2. Place standard order (priority=10)\n3. Go to Orders page\n4. Check order of cards",
     "isPremium: true",
     "Premium order (priority=1) listed above standard (priority=10)","","","P1","Regression"],

    ["TC_O_03","Orders","New order status shows PENDING",
     "Order just placed",
     "1. Place order\n2. Navigate to /orders\n3. Check status badge",
     "N/A",
     "Status badge shows 'PENDING'","","","P2","Regression"],

    ["TC_O_04","Orders","Order contains correct book items and total",
     "Order placed with specific books",
     "1. Add book (₹599) to cart\n2. Place order\n3. Open order on Orders page\n4. Verify item name and total",
     "Book: The Pragmatic Programmer ₹599",
     "Order shows correct book name, qty and total amount","","","P1","Regression"],

    ["TC_O_05","Orders","Multiple orders are all listed",
     "User has placed 3+ orders",
     "1. Place 3 separate orders\n2. Navigate to /orders\n3. Count order cards",
     "N/A",
     "All 3 orders visible on the page","","","P2","Regression"],

    ["TC_O_06","Orders","Orders sorted by priority — premium always first",
     "Multiple orders with different priorities",
     "1. Place standard order\n2. Place premium order\n3. Go to /orders\n4. Verify sort order",
     "Standard priority=10\nPremium priority=1",
     "Orders sorted ascending by priority — premium (1) before standard (10)","","","P1","Regression"],

    ["TC_O_07","Orders","Empty orders page shows empty state",
     "User with no orders placed",
     "1. Log in as new user with no orders\n2. Navigate to /orders",
     "No prior orders",
     "Empty state message shown — no errors, no blank white screen","","","P2","Functional"],
]

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 5 — API Tests  (5 test cases)
# ═══════════════════════════════════════════════════════════════════════════════
api_tcs = [
    ["TC_A_01","API","GET /books returns 200 with 12 books",
     "Valid anon key configured in ApiTest.java",
     "1. Send GET /rest/v1/books?select=*\n2. Include apikey header\n3. Check status code and body",
     "apikey: sb_publishable_5pi0OY09IbihgQliazw5ug_G5ha_SxJ",
     "HTTP 200, JSON array with 12 book objects",
     "HTTP 200 returned — 12 books in response","PASS","P1","API"],

    ["TC_A_02","API","GET /coupons returns 200 with 4 coupons",
     "Valid anon key available",
     "1. Send GET /rest/v1/coupons?select=*\n2. Include apikey header",
     "apikey: sb_publishable_5pi0OY09...",
     "HTTP 200, JSON array with 4 coupon objects (BOOK30/80/150/300)",
     "HTTP 200 returned — 4 coupons in response","PASS","P1","API"],

    ["TC_A_03","API","GET /books filtered by category=Fiction returns 200",
     "Valid anon key available",
     "1. Send GET /rest/v1/books?select=*&category=eq.Fiction",
     "category=Fiction",
     "HTTP 200, only Fiction category books in response",
     "HTTP 200 returned — Fiction books filtered correctly","PASS","P2","API"],

    ["TC_A_04","API","GET /books sorted by price ascending returns 200",
     "Valid anon key available",
     "1. Send GET /rest/v1/books?select=*&order=price.asc",
     "order=price.asc",
     "HTTP 200, books ordered by price lowest first",
     "HTTP 200 returned — books ordered by price ascending","PASS","P2","API"],

    ["TC_A_05","API","GET /carts without JWT — RLS returns 200 empty array",
     "No Authorization header sent",
     "1. Send GET /rest/v1/carts?select=*\n2. Include only apikey header\n3. No Authorization header",
     "Authorization header: missing",
     "HTTP 200 — Supabase RLS returns empty array (no data leak for anon)",
     "HTTP 200 with empty array — no cart data exposed to anon user","PASS","P1","API"],
]

# ═══════════════════════════════════════════════════════════════════════════════
# Build workbook — total = 14+12+12+7+5 = 50
# ═══════════════════════════════════════════════════════════════════════════════
make_sheet(wb, "Login Tests",   "2E75B6", login_tcs,   first=True)
make_sheet(wb, "Catalog Tests", "375623", catalog_tcs)
make_sheet(wb, "Cart Tests",    "7030A0", cart_tcs)
make_sheet(wb, "Orders Tests",  "C55A11", orders_tcs)
make_sheet(wb, "API Tests",     "1F3864", api_tcs)

total = len(login_tcs) + len(catalog_tcs) + len(cart_tcs) + len(orders_tcs) + len(api_tcs)
out = "/home/varad/projects/E-commerce/selenium-tests/BookSphere_TestCases.xlsx"
wb.save(out)
print(f"Saved: {out}  |  Total test cases: {total}")
